#!/usr/local/python37/bin/python3.7
# evaluate the averaged NMR chemical shifts
import os
import re
from openpyxl import Workbook

def extract_energy_from_log(log_file, keyword):
    with open(log_file, 'r') as file:
        lines = file.readlines()

    for i, line in enumerate(lines):
        if keyword in line:
            energy_line = lines[i].strip()
            energy_match = re.search(r'-?\d+\.\d+', energy_line)
            if energy_match:
                return float(energy_match.group())

    return None

def extract_boltzmann_weights_from_txt(txt_file):
    weights = []
    with open(txt_file, 'r') as file:
        lines = file.readlines()

    for line in lines:
        if "Boltzmann weight=" in line:
            weight_line = line.strip()
            weight_match = re.search(r'Boltzmann weight=(.{10})', weight_line)
            if weight_match:
                weights.append(weight_match.group(1).strip())

    return weights

def main():
    # 获取当前目录下所有 *.log 文件
    log_files = [file for file in os.listdir() if file.endswith(".log")]

    if not log_files:
        print("当前目录下找不到 *.log 文件。")
        return

    # 创建一个新的Excel工作簿
    wb = Workbook()
    ws = wb.active

    # 添加表头
    ws['A1'] = '文件名'
    ws['B1'] = 'FINAL SINGLE POINT ENERGY'
    ws['C1'] = '文件名'
    ws['D1'] = 'Thermal correction to Gibbs Free Energy'
    ws['E1'] = 'Boltzmann weight'
    ws['F1'] = '文件名'

    # 提取数据并写入Excel
    row_num = 2
    for log_file in log_files:
        energy_value = extract_energy_from_log(log_file, "FINAL SINGLE POINT ENERGY")
        gibbs_energy_value = extract_energy_from_log(log_file, "Thermal correction to Gibbs Free Energy")

        if energy_value is not None:
            ws.cell(row=row_num, column=1, value=log_file)
            ws.cell(row=row_num, column=2, value=energy_value)
            row_num += 1

        if gibbs_energy_value is not None:
            ws.cell(row=row_num, column=3, value=log_file)
            ws.cell(row=row_num, column=4, value=gibbs_energy_value)
            row_num += 1

    # 遍历 *_08_sher01rate.txt 文件
    sher_files = [file for file in os.listdir() if file.endswith("_08_sher01rate.txt")]

    for sher_file in sher_files:
        weights = extract_boltzmann_weights_from_txt(sher_file)

        for weight_value in weights:
            ws.cell(row=row_num, column=5, value=weight_value)
            ws.cell(row=row_num, column=6, value=sher_file)
            row_num += 1

    # 保存Excel文件
    wb.save('energy_weights_data.xlsx')
    print("数据已成功写入 energy_weights_data.xlsx 文件。")

if __name__ == "__main__":
    main()
